#!/usr/bin/env python
# -*- coding:utf8 -*-
# !/usr/bin/env python
# -*- coding:utf8 -*-
from xml.dom import minidom
from xml.dom.minidom import parse
import xml.etree.ElementTree as ET
import time

import operator as op

# 根据中文的翻译文件，生成英文的
import openpyxl
from openpyxl import load_workbook

workbook = load_workbook("C:/Users/wangtao/Desktop/FiiiPOS国际化.xlsx")
sheel = workbook["062018"]
dirctZh = {}
i = 1
for row in sheel.rows:
    i += 1
    # 第一列
    colKey = sheel.cell(row=i, column=1).value
    if colKey is None:
        break
    # 第三列
    colValue = sheel.cell(row=i, column=3).value
    dirctZh[colKey] = colValue

print(dirctZh)  # 打印结果

out_file = open("C:/Users/wangtao/Desktop/strings-res.xml", 'w+', encoding='utf-8')

out_file.write("<resources>\n")
for key in dirctZh:
    out_file.write("<string name=\"" + key + "\">" + dirctZh.get(key) + "</string>\n")
out_file.write("</resources>")
